from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import sqlite3
import qrcode
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
from pyzbar.pyzbar import decode
from PIL import Image
import socket
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "secret_key"

# Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

UPLOAD_FOLDER = 'gestion_ely/static/uploads/'
EXCEL_FOLDER = 'gestion_ely/static/excel/'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['EXCEL_FOLDER'] = EXCEL_FOLDER

# Crear carpetas si no existen
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXCEL_FOLDER, exist_ok=True)

class User(UserMixin):
    def __init__(self, id, username, password):
        self.id = id
        self.username = username
        self.password = password

@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()
    c.execute("SELECT id, username, password FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    conn.close()
    if user:
        return User(id=user[0], username=user[1], password=user[2])
    return None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def crear_tablas():
    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()
    c.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL,
        password TEXT NOT NULL
    )
    ''')
    c.execute('''
    CREATE TABLE IF NOT EXISTS inventario (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        referencia TEXT NOT NULL,
        peso REAL NOT NULL,
        fecha_entrada TEXT NOT NULL,
        img_path TEXT NOT NULL
    )
    ''')
    c.execute('''
    CREATE TABLE IF NOT EXISTS camiones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        referencia TEXT NOT NULL,
        matricula TEXT,
        empresa TEXT,
        fecha_salida TEXT,
        peso_total REAL NOT NULL,
        estado TEXT NOT NULL DEFAULT 'Preparado'
    )
    ''')
    c.execute('''
    CREATE TABLE IF NOT EXISTS camiones_inventario (
        camion_id INTEGER NOT NULL,
        inventario_id INTEGER NOT NULL,
        FOREIGN KEY (camion_id) REFERENCES camiones(id),
        FOREIGN KEY (inventario_id) REFERENCES inventario(id)
    )
    ''')
    conn.commit()
    conn.close()

def crear_usuario_predeterminado():
    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE username = 'usuarioely'")
    user = c.fetchone()
    if not user:
        hashed_password = generate_password_hash("usuarioely1", method='pbkdf2:sha256')
        c.execute("INSERT INTO users (username, password) VALUES (?, ?)", ('usuarioely', hashed_password))
        conn.commit()
    conn.close()

crear_tablas()
crear_usuario_predeterminado()

def generar_etiqueta(referencia, peso, fecha_entrada):
    datos = f'Referencia: {referencia}, Peso: {peso}, Fecha Entrada: {fecha_entrada}'
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(datos)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    
    os.makedirs('gestion_ely/static/etiquetas', exist_ok=True)

    fecha_entrada_segura = fecha_entrada.replace('/', '-')
    img_path = os.path.join('gestion_ely/static/etiquetas', f'{referencia}_{peso}_{fecha_entrada_segura}.png')
    img.save(img_path)
    return img_path

def actualizar_excel():
    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()

    # Actualizar inventario
    c.execute("SELECT referencia, peso, fecha_entrada, img_path FROM inventario")
    datos_inventario = c.fetchall()
    
    # Actualizar camiones en preparación
    c.execute("SELECT referencia, peso_total FROM camiones WHERE estado = 'Preparado'")
    datos_preparacion = c.fetchall()
    
    # Actualizar camiones de salida
    c.execute("SELECT referencia, matricula, empresa, fecha_salida, peso_total FROM camiones WHERE estado = 'Salida'")
    datos_salida = c.fetchall()

    conn.close()
    
    excel_path = os.path.join(app.config['EXCEL_FOLDER'], "gestion_ely.xlsx")
    if os.path.exists(excel_path):
        libro_excel = load_workbook(excel_path)
    else:
        libro_excel = Workbook()
        libro_excel.create_sheet("Inventario")
        libro_excel.create_sheet("Preparación de Camiones")
        libro_excel.create_sheet("Salida de Camiones")
        del libro_excel['Sheet']

    # Actualizar hoja de Inventario
    hoja_inventario = libro_excel["Inventario"]
    hoja_inventario.delete_rows(2, hoja_inventario.max_row)
    encabezados_inventario = ["Referencia", "Peso", "Fecha de Entrada", "Imagen QR"]
    hoja_inventario.append(encabezados_inventario)
    for registro in datos_inventario:
        hoja_inventario.append(registro)
    
    # Actualizar hoja de Preparación de Camiones
    hoja_preparacion = libro_excel["Preparación de Camiones"]
    hoja_preparacion.delete_rows(2, hoja_preparacion.max_row)
    encabezados_preparacion = ["Referencia", "Peso Total"]
    hoja_preparacion.append(encabezados_preparacion)
    for registro in datos_preparacion:
        hoja_preparacion.append(registro)
    
    # Actualizar hoja de Salida de Camiones
    hoja_salida = libro_excel["Salida de Camiones"]
    hoja_salida.delete_rows(2, hoja_salida.max_row)
    encabezados_salida = ["Referencia", "Matrícula", "Empresa", "Fecha de Salida", "Peso Total"]
    hoja_salida.append(encabezados_salida)
    for registro in datos_salida:
        hoja_salida.append(registro)

    libro_excel.save(excel_path)

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = sqlite3.connect('inventario.db')
        c = conn.cursor()
        c.execute("SELECT id, username, password FROM users WHERE username = ?", (username,))
        user = c.fetchone()
        conn.close()
        if user and check_password_hash(user[2], password):
            login_user(User(id=user[0], username=user[1], password=user[2]))
            return redirect(url_for('index'))
        else:
            flash('Nombre de usuario o contraseña incorrectos')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/llegada', methods=['GET', 'POST'])
@login_required
def llegada():
    if request.method == 'POST':
        referencia = request.form['referencia']
        peso = request.form['peso']
        fecha_entrada = request.form['fecha_entrada']
        if not referencia or not peso or not fecha_entrada:
            flash('Por favor, complete todos los campos')
            return redirect(url_for('llegada'))
        try:
            peso = float(peso)
        except ValueError:
            flash('Peso debe ser un número válido')
            return redirect(url_for('llegada'))
        img_path = generar_etiqueta(referencia, peso, fecha_entrada)
        conn = sqlite3.connect('inventario.db')
        c = conn.cursor()
        c.execute("INSERT INTO inventario (referencia, peso, fecha_entrada, img_path) VALUES (?, ?, ?, ?)",
                  (referencia, peso, fecha_entrada, img_path))
        conn.commit()
        conn.close()
        actualizar_excel()
        flash('Producto agregado con éxito.')
        return redirect(url_for('llegada'))
    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()
    c.execute("SELECT referencia, peso, fecha_entrada FROM inventario ORDER BY fecha_entrada DESC")
    registros = c.fetchall()
    conn.close()
    return render_template('llegada.html', registros=registros)

@app.route('/preparacion', methods=['GET', 'POST'])
@login_required
def preparacion():
    if 'camion_id' not in session:
        return redirect(url_for('crear_camion'))
    
    if request.method == 'POST':
        archivo_qr = request.files['qr']
        if archivo_qr and allowed_file(archivo_qr.filename):
            filename = secure_filename(archivo_qr.filename)
            qr_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            archivo_qr.save(qr_path)
            img = Image.open(qr_path)
            img.load()
            codigo_qr = decode(img)
            if codigo_qr:
                datos = codigo_qr[0].data.decode("utf-8")
                referencia, peso, fecha_entrada = [x.split(': ')[1] for x in datos.split(', ')]
                peso = float(peso)
                conn = sqlite3.connect('inventario.db')
                c = conn.cursor()
                c.execute("SELECT id FROM inventario WHERE referencia = ? AND peso = ? AND fecha_entrada = ?",
                          (referencia, peso, fecha_entrada))
                inventario_id = c.fetchone()
                if inventario_id:
                    inventario_id = inventario_id[0]
                    c.execute("INSERT INTO camiones_inventario (camion_id, inventario_id) VALUES (?, ?)",
                              (session['camion_id'], inventario_id))
                    c.execute("UPDATE camiones SET peso_total = peso_total + ? WHERE id = ?",
                              (peso, session['camion_id']))
                    conn.commit()
                    flash(f'Producto {referencia} agregado al camión')
                else:
                    flash('Producto no encontrado en el inventario')
                conn.close()
            else:
                flash('No se pudo escanear el código QR')
        else:
            flash('Archivo no permitido o no se ha subido ningún archivo')
        return redirect(url_for('preparacion'))

    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()
    c.execute("SELECT id, referencia, peso_total FROM camiones WHERE estado = 'Preparado'")
    camiones_preparados = c.fetchall()
    conn.close()
    return render_template('preparacion.html', camiones_preparados=camiones_preparados)

@app.route('/crear_camion', methods=['GET', 'POST'])
@login_required
def crear_camion():
    referencia_camion = f"CAMION-{int(datetime.now().timestamp())}"
    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()
    c.execute("INSERT INTO camiones (referencia, matricula, empresa, fecha_salida, peso_total, estado) VALUES (?, ?, ?, ?, ?, ?)",
              (referencia_camion, 'N/A', 'N/A', 'N/A', 0, 'Preparado'))
    camion_id = c.lastrowid
    conn.commit()
    conn.close()
    session['camion_id'] = camion_id
    return redirect(url_for('preparacion'))

@app.route('/finalizar_camion', methods=['POST'])
@login_required
def finalizar_camion():
    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()
    c.execute("UPDATE camiones SET estado = 'Salida' WHERE id = ?", (session['camion_id'],))
    conn.commit()
    conn.close()
    actualizar_excel()
    session.pop('camion_id', None)
    flash('Camión preparado con éxito.')
    return redirect(url_for('index'))

@app.route('/salida', methods=['GET', 'POST'])
@login_required
def salida():
    if request.method == 'POST':
        camion_id = request.form['camion_id']
        matricula = request.form['matricula']
        empresa = request.form['empresa']
        fecha_salida = request.form['fecha_salida']
        conn = sqlite3.connect('inventario.db')
        c = conn.cursor()
        c.execute("SELECT referencia, peso, fecha_entrada FROM inventario i JOIN camiones_inventario ci ON i.id = ci.inventario_id WHERE ci.camion_id = ?", (camion_id,))
        productos = c.fetchall()
        c.execute("UPDATE camiones SET estado = 'Salida', matricula = ?, empresa = ?, fecha_salida = ? WHERE id = ?", (matricula, empresa, fecha_salida, camion_id))
        conn.commit()
        conn.close()
        
        actualizar_excel()
        
        flash('Camión salido registrado con éxito.')
        return redirect(url_for('salida'))
    
    conn = sqlite3.connect('inventario.db')
    c = conn.cursor()
    c.execute("SELECT id, referencia, peso_total FROM camiones WHERE estado = 'Preparado'")
    camiones_preparados = c.fetchall()
    conn.close()
    return render_template('salida.html', camiones_preparados=camiones_preparados)

@app.route('/download/<filename>')
@login_required
def download_file(filename):
    return send_from_directory(app.config['EXCEL_FOLDER'], filename, as_attachment=True)

def get_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = '127.0.0.1'
    finally:
        s.close()
    return ip

if __name__ == '__main__':
    ip = get_ip()
    port = 5007
    print(f'Servidor corriendo en http://{ip}:{port}')
    app.run(debug=True, host='0.0.0.0', port=port)
